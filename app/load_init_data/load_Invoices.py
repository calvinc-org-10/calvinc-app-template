from typing import Any, Dict, List

import re
from openpyxl import load_workbook
from datetime import date, datetime
from incShip.models import HBL, ShippingForms, Containers, Invoices, PO, references
from incShip.models import Companies, FreightTypes, Organizations

from PySide6.QtCore import (QCoreApplication, QDate, QDateTime, QLocale,
    QMetaObject, QObject, QPoint, QRect,
    QSize, QTime, QUrl, Qt)
from PySide6.QtGui import (QBrush, QColor, QConicalGradient, QCursor,
    QFont, QFontDatabase, QGradient, QIcon,
    QImage, QKeySequence, QLinearGradient, QPainter,
    QPalette, QPixmap, QRadialGradient, QTransform)
from PySide6.QtWidgets import (QApplication, QWidget, 
    QGridLayout, QSpacerItem,
    QListView, QListWidget, QColumnView, QListWidgetItem, QTableView, QHeaderView,
    QComboBox, QDateEdit, QFrame,
    QLabel, QLineEdit, 
    QPushButton,
    QSizePolicy, QTabWidget, QTextEdit,
    )

from forms import std_popdialogsize 

from cMenu.utils import cQFmNameLabel


_DEFAULTDATE = date.today()

class LoadInvoices(QWidget):

    lblStatus = QLabel()

    def __init__(self, parent = None):
        super().__init__(parent)
        if not self.objectName():
            self.setObjectName(u"Form")
        self.resize(2*std_popdialogsize)

        self.lblFormName = cQFmNameLabel(parent=self)
        wdgt = self.lblFormName
        wdgt.setObjectName(u"lblFormName")
        wdgt.setGeometry(QRect(150, 10, 400, 74))
        
        wdgt = self.lblStatus
        wdgt.setParent(self)
        wdgt.setWordWrap(True)
        wdgt.setGeometry(10, 30, 600, 400)

        self.retranslateUi()

        self.show()
        self.load_recs()
    # __init__

    def retranslateUi(self):
        self.setWindowTitle(QCoreApplication.translate("Form", "Load Invoices from SprSht", None))
        self.lblFormName.setText(QCoreApplication.translate("Form", "Load Invoices from SprSht", None))
   # retranslateUi

    def load_recs(self):
        nRowsAdded = 0
        SprshtRowNum = 0
        fName = f'W:\\Logistics\\Invoices\\SmOff Expense Forms_Exped-JUSDA.xlsx'

        wb = load_workbook(filename=fName, read_only=True)
        ws = wb.active

        _SStName_InvNum = 'Invoice ID'
        _TblName_InvNum = 'InvoiceNumber'
        colmnNames = ws[1]
        colmnMap = {_TblName_InvNum: None, }    # start with columns that MUST be in the SpSht
        SSName_TableName_map = {
                _SStName_InvNum : _TblName_InvNum, 
                'Form ID': 'id_SmOffFormNum',
                'Invoice Date': 'InvoiceDate',
                'Ground Total': 'InvoiceAmount',
                'Business Unit': 'ORG',
                'Vendor Code': 'Company',
                'Item Status': 'SmOffStatus',
                #    class SmOffStatusCodes(models.IntegerChoices):
                #        NOTENT   = 000, 'NOT ENTRD'
                #        DRAFT    = 100, 'DRAFT'   
                #        PENDING  = 200, 'PENDING'
                #        APPROVED = 900, 'APPROVED'
                'Tracking No': 'HBL',
                'Memo': 'notes', # extract POs
                'Item Description': '*FrType', 
                }
        # map column names to column numbers
        for col in colmnNames:
            if col.value in SSName_TableName_map:
                colkey = SSName_TableName_map[col.value]
                # has this col.value already been mapped?
                if (colkey in colmnMap and colmnMap[colkey] is not None):
                    # yes, that's a problem
                    statecode = 'fatalerr'
                    statetext = f'SAP Spreadsheet has bad header row - More than one column named {col.value}.  See Calvin to fix this.'
                    result = 'FAIL - bad spreadsheet'
                    self.lblStatus.setText(f'{statecode}\n{statetext}\n{result}')
                    wb.close()
                    return
                else:
                    colmnMap[colkey] = col.column - 1
                # endif previously mapped
            #endif col.value in SAP_SSName_TableName_map
        #endfor col in SAPcolmnNames
        if (colmnMap[_TblName_InvNum] is None):
            statecode = 'fatalerr'
            statetext = f'SAP Spreadsheet has bad header row - no {_SStName_InvNum} column.  See Calvin to fix this.'
            result = 'FAIL - bad spreadsheet'
            self.lblStatus.setText(f'{statecode}\n{statetext}\n{result}')
            wb.close()
            return 

        numrows = ws.max_row
        numrow98 = .98*numrows
        numrow90 = .90*numrows
        SprshtRowNum = 1
        announceInterval = 100
        
        for row in ws.iter_rows(min_row=SprshtRowNum+1, values_only=True):
            SprshtRowNum += 1
            
            if SprshtRowNum >= numrow98:
                announceInterval = 1
            elif SprshtRowNum >= numrow90:
                announceInterval = 10
                
            if SprshtRowNum % announceInterval == 0:
                statecode = 'rdng-sprsht'
                statetext = f'Reading Spreadsheet ... record {SprshtRowNum} of {numrows}'
                self.lblStatus.setText(f'{statecode}\t{statetext}\n{row}')
                print(f'{statecode}\t{statetext}\n{row}')
                self.lblStatus.repaint()

            if row[colmnMap[_TblName_InvNum]]==None: InvcNum = ''
            else: InvcNum = row[colmnMap[_TblName_InvNum]]
            if len(str(InvcNum)):
                InvcRec, newrec = Invoices.objects.get_or_create(InvoiceNumber=InvcNum,
                    defaults={
                        'InvoiceDate': _DEFAULTDATE,
                        'InvoiceAmount': 0,
                        'HBL': HBL.objects.first(),
                        'SmOffStatus': Invoices.SmOffStatusCodes.PENDING,
                        'inv_downloaded': True,
                        'verifiedForFii': True,
                        })
                CompanyRec = Companies.objects.filter(SmOffVendorID=row[colmnMap['Company']]).first() 
                FrType = FreightTypes.objects.first()
                if FreightTypes.objects.filter(FreightType=row[colmnMap['*FrType']]).exists():
                    FrType = FreightTypes.objects.filter(FreightType=row[colmnMap['*FrType']]).first()
                HBLRec, newrec = HBL.objects.get_or_create(HBLNumber=row[colmnMap['HBL']],
                    defaults= {
                        'Company': CompanyRec,
                        'FreightType': FrType,
                    })
                for fldName, colNum in colmnMap.items():
                    if fldName[0] == '*':       # dummy field, skip it
                        continue
                    if fldName == _TblName_InvNum:
                        # no transform needed - correctly set in get_or_create above
                        pass
                    if fldName == 'id_SmOffFormNum':
                        setattr(InvcRec, fldName, row[colNum])
                    if fldName == 'InvoiceDate':
                        # D = isDate(row[colNum])
                        D = datetime.strptime(row[colNum], "%Y-%m-%d %H:%M:%S").date()
                        if D and getattr(InvcRec,fldName) == _DEFAULTDATE: setattr(InvcRec, fldName, D)
                    if fldName == 'InvoiceAmount':
                        # Amt = isNumeric(row[colNum])
                        Amt = row[colNum]
                        if Amt and getattr(InvcRec, fldName) == 0: setattr(InvcRec, fldName, Amt)
                    Orgrec = Organizations.objects.filter(ApplyingBusinesUnit__contains='L6').first()
                    if fldName == 'ORG':
                        if re.search('L10', row[colNum], re.IGNORECASE):
                            Orgrec = Organizations.objects.filter(ApplyingBusinesUnit__contains='L10').first()
                    if fldName == 'Company':
                        if row[colNum]: setattr(InvcRec, fldName, CompanyRec )
                    if fldName == 'FreightType':
                        setattr(InvcRec, fldName, FrType)
                    if fldName == 'SmOffStatus':
                        if row[colNum].lower() == 'approved': setattr(InvcRec, fldName, Invoices.SmOffStatusCodes.APPROVED)
                    if fldName == 'HBL':
                        if HBLRec: setattr(InvcRec, fldName, HBLRec)
                    if fldName == 'notes' and HBLRec: # extract POs
                        noteStr = str(row[colNum]) if row[colNum] is not None else ''
                        setattr(InvcRec, fldName, noteStr)
                        POinInvoice = re.findall('4[0-9]{9}',noteStr)
                        for Px in POinInvoice:
                            PORec, crFlag = PO.objects.get_or_create(PONumber=Px, defaults={'org': Orgrec})
                            HBLRec.POList.add(PORec)
                InvcRec.save()
                nRowsAdded += 1
                # endif MatlRec
            # endif len(str(MatlNum))
        #endfor row in ws.iter_rows

        wb.close()
        self.lblStatus.setText('Invoices pulled from spreadsheet')

