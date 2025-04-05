from typing import Any, Dict, List

import re
from openpyxl import load_workbook
from datetime import date
from django.http import HttpResponse
from incShip.models import HBL, ShippingForms, Containers, Invoices, references
from incShip.models import Companies, FreightTypes

from PySide6.QtCore import (QCoreApplication, QDate, QDateTime, QLocale,
    QMetaObject, QObject, QPoint, QRect,
    QSize, QTime, QUrl, Qt)
from PySide6.QtGui import (QBrush, QColor, QConicalGradient, QCursor,
    QFont, QFontDatabase, QGradient, QIcon,
    QImage, QKeySequence, QLinearGradient, QPainter,
    QPalette, QPixmap, QRadialGradient, QTransform)
from PySide6.QtWidgets import (QApplication, QWidget, 
    QGridLayout, QSpacerItem,
    QListView, QListWidget, QColumnView, QListWidgetItem, QTableView, QHeaderView,
    QComboBox, QDateEdit, QFrame,
    QLabel, QLineEdit, 
    QPushButton,
    QSizePolicy, QTabWidget, QTextEdit,
    )

from forms import std_popdialogsize

from cMenu.utils import cQFmNameLabel

class LoadHBL(QWidget):
    _linkedTables = ['ShippingForms', 'PO', 'Invoices', 'Containers', 'reference_ties']

    lblStatus = QLabel()

    def __init__(self, parent:QWidget = None):
        super().__init__(parent)

        if not self.objectName():
            self.setObjectName(u"Form")
        self.resize(2*std_popdialogsize)

        self.lblFormName = cQFmNameLabel(parent=self)
        wdgt = self.lblFormName
        wdgt.setObjectName(u"lblFormName")
        wdgt.setGeometry(QRect(200, 10, 319, 74))
        
        wdgt = self.lblStatus
        wdgt.setParent(self)
        wdgt.setWordWrap(True)
        wdgt.setGeometry(10, 30, 600, 400)

        self.retranslateUi()

        self.show()
        self.load_recs()
    # __init__

    def retranslateUi(self):
        self.setWindowTitle(QCoreApplication.translate("Form", "Load HBL from SprSht", None))
        self.lblFormName.setText(QCoreApplication.translate("Form", "Load HBL from SprSht", None))
   # retranslateUi

    def load_recs(self):
        fName = f'W:\\Logistics\\Invoices\\Invoices.xlsx'   #TODO: put this on form
        shtName = 'HBL heads-ups'                           #TODO: put this on form
        wb = load_workbook(filename=fName, read_only=True)
        ws = wb[shtName]
        
        process_flag = False
        import_spsht_prfx='SpSht.Import'
        sentinel_col, sentinel_flag = ('LstDigChk', 'RESTART')
        
        colmnNames = ws[1]
        colmnMap = { }    # start with columns that MUST be in the SpSht
        SSName_TableName_map = {
            sentinel_col: None ,         # used only to find the sentinel
            'HBL': 'HBLNumber', 
            'sh fm': 'ShippingForm',              # mod/create ShippingForm
            'container num': 'Container',      # mod/create a Container
            'mode': 'FreightType', 
            'origin': 'Origin', 
            'notes': 'notes',           # in the refs table
            'Pickup dt': 'PickupDt', 
            'ETA ORD or MKE': 'ETA', 
            'Deliv appt': 'DelivAppt',  # in Containers table
            'LFD': 'LFD',               # in HBL AND CXontainers
            'orig invc': 'InvoiceNumber',  # mod/create Invoices
            'orig SmOff': 'id_SmOffFormNum',  # mod/create Invoices
            'Company': 'Company',
            }
        # map column names to column numbers
        for col in colmnNames:
            if col.value in SSName_TableName_map:
                colkey = col.value # SSName_TableName_map[col.value] - not used - I do the spSht -> Table mappings manually here
                # has this col.value already been mapped?
                if (colkey in colmnMap and colmnMap[colkey] is not None):
                    # yes, that's a problem
                    statecode = 'fatalerr'
                    statetext = f'SAP Spreadsheet has bad header row - More than one column named {col.value}.  See Calvin to fix this.'
                    result = 'FAIL - bad spreadsheet'
                    self.lblStatus.setText(f'{statecode}\n{statetext}\n{result}')
                    wb.close()
                    return
                else:
                    colmnMap[colkey] = col.column - 1
                # endif previously mapped
            #endif col.value in SAP_SSName_TableName_map
        #endfor col in SAPcolmnNames
        
        numrows = ws.max_row
        numrow98 = .98 * numrows
        numrow90 = .90 * numrows
        SprshtRowNum = 1
        announceInterval = 100
        
        Company_dflt =  Companies.objects.get(pk=1)    # not always true, but I'll fix individual cases manually
        
        for row in ws.iter_rows(min_row=SprshtRowNum+1, values_only=True):
            SprshtRowNum += 1
            
            if not process_flag or row[colmnMap[sentinel_col]] == sentinel_flag:
                process_flag = True
                continue
            # endif process_flag
            

            HBLNum = row[colmnMap['HBL']]
            ShFm = row[colmnMap['sh fm']]
            ContNum = row[colmnMap['container num']]
            InvNum = row[colmnMap['orig invc']]
            print(SprshtRowNum)
            Company = Companies.objects.get(CompanyName__istartswith=row[colmnMap['Company']]) if row[colmnMap['Company']] else Company_dflt
            if InvNum:
                if re.match(r'bill.*', InvNum): InvNum = None

            PkupDt = row[colmnMap['Pickup dt']]
            ETA = row[colmnMap['ETA ORD or MKE']]
            DelivAppt = row[colmnMap['Deliv appt']]
            LFD = row[colmnMap['LFD']]
            #create records
            HBLrec = None
            if HBLNum:
                HBLrec, creatFlag =  HBL.objects.get_or_create(HBLNumber = HBLNum, defaults={'notes': ''})
                HBLrec.Company = Company
                if PkupDt: HBLrec.PickupDt = PkupDt
                if ETA: HBLrec.ETA = ETA
                if LFD: HBLrec.LFD = LFD
                FType = FreightTypes.objects.first()
                v = row[colmnMap['mode']]
                if v:
                    if re.search(r'air', v,re.IGNORECASE): FType = FreightTypes.objects.filter(FreightType__icontains='air').first()
                    if re.search(r'ocean', v, re.IGNORECASE): FType = FreightTypes.objects.filter(FreightType__icontains='ocean').first()
                HBLrec.FreightType = FType
                HBLrec.save()
            ShFmrec = None
            if ShFm:
                ShFmrec, creatFlag = ShippingForms.objects.get_or_create(id_SmOffFormNum = ShFm, defaults={'notes': ''})
            Contrec = None
            if ContNum:
                Contrec, creatFlag = Containers.objects.get_or_create(ContainerNumber = ContNum, defaults={'notes': ''})
                if DelivAppt: Contrec.DelivAppt = DelivAppt
                if LFD: Contrec.LFD = LFD
                Contrec.save()
            Invrec = None
            if InvNum and HBLNum:
                Invrec, creatFlag = Invoices.objects.get_or_create(InvoiceNumber=InvNum,
                    defaults= {
                        'HBL':HBLrec,
                        'SmOffStatus': Invoices.SmOffStatusCodes.PENDING,
                        'InvoiceDate': date.today(),
                        'InvoiceAmount': 0,
                        'notes': '',
                    }
                )
                Invrec.verifiedForFii = True
                Invrec.inv_downloaded = True
                if row[colmnMap['orig SmOff']] and not Invrec.id_SmOffFormNum:
                    Invrec.id_SmOffFormNum = row[colmnMap['orig SmOff']]
                Invrec.save()
            
            if HBLrec and ShFmrec: 
                HBLrec.ShippingForms.add(ShFmrec)
            if HBLrec and Contrec:
                Contrec.HBL = HBLrec
                Contrec.save()
            noterec = None
            if row[colmnMap['notes']]:
                noterec, creatFlag = references.objects.get_or_create(
                    refName = f'{import_spsht_prfx}.{SprshtRowNum}',
                    defaults={
                        'notes': row[colmnMap['notes']]
                    }
                )
                if HBLrec: noterec.HBL = HBLrec
                if ShFmrec: noterec.ShippingForm = ShFmrec
                if Contrec: noterec.Container = Contrec
                if Invrec: noterec.Invoice = Invrec
                noterec.save()

            self.lblStatus.setText(f'{SprshtRowNum}, {row}, {HBLNum}, {ShFm}, {ContNum}, {InvNum} processed')

        self.lblStatus.setText('Data pulled from spreadsheet')
    # load_recs

