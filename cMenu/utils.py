from typing import (Dict, List, Any, Tuple, NamedTuple )
from collections import (namedtuple, )
from collections.abc import (Callable, )

from PySide6.QtCore import (QCoreApplication, 
    QDate, QDateTime, QLocale,
    QMetaObject, QObject, QPoint, QRect,
    QSize, QTime, QUrl, Qt,
    Signal, Slot, 
    QStringListModel, QAbstractTableModel, QAbstractListModel, QModelIndex,
    )
from PySide6.QtGui import (QBrush, QColor, QConicalGradient, QCursor,
    QFont, 
    )
from PySide6.QtSql import (QSqlDatabase, QSqlQueryModel, QSqlRelationalTableModel, QSqlTableModel, QSqlRecord, )
from PySide6.QtWidgets import (QApplication, QWidget, QGridLayout, QHBoxLayout, QVBoxLayout,
    QScrollArea, QCompleter, 
    QDialog, QMessageBox, QDialogButtonBox, 
    QPushButton, QLabel, QFrame, QLineEdit, QTextEdit, QPlainTextEdit, QDateEdit, 
    QCheckBox, QComboBox, 
    QSizePolicy, 
    )
from PySide6.QtSvgWidgets import QSvgWidget

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, fills, colors
from openpyxl.utils.datetime import from_excel, WINDOWS_EPOCH


# standard window and related sizes
# copied from main app's forms module
std_windowsize = QSize(1120,720)
std_popdialogsize=QSize(400,300)

ExcelWorkbook_fileext = ".XLSX"

def pleaseWriteMe(parent, addlmessage):
    msg = QMessageBox(parent)
    msg.setWindowTitle('Please Write Me')
    msg.setIcon(QMessageBox.Icon.Warning)
    msg.setStandardButtons(QMessageBox.StandardButton.Ok)
    msg.setText(f'Calvin needs to get up off his butt and write some code\n{addlmessage}')
    msg.open()

def areYouSure(parent:QWidget, title:str, 
        areYouSureQuestion:str, 
        answerChoices:QMessageBox.StandardButton = QMessageBox.StandardButton.Yes|QMessageBox.StandardButton.No,
        dfltAnswer:QMessageBox.StandardButton = QMessageBox.StandardButton.No,
        ) -> QMessageBox.StandardButton:
    ret = QMessageBox.question(parent, title,
        areYouSureQuestion, answerChoices, dfltAnswer)
    return ret

class cDataList(QLineEdit):
    """
    A LineEdit box that acts like an HTML DataList
    Choice matches are choices which contain the input string, case-insensitive
    
    caller should connect the editingFinished signal to a slot which is aware of the cDataList
        and is in scope to call cDataList.selectedItem
    ex: self.testdatalist.editingFinished.connect(self.showHBLChosen)

    Args:
        choices:Dict[Any, str], {key: 'value to display/lookup'}
        initval:str = '', (not currently implemented)
        parent:QWidget=None

    def selectedItem(self):
        returns the following dictionary: 
        return {'keys': [key for key, t in self.choices.items() if t==self.text()], 'text': self.text()}
        (all keys matching the text input)
    """
    def __init__(self, choices:Dict[Any, str], initval:str = '', parent:QWidget=None):
        super().__init__(initval, parent)

        self.choices = choices
        
        self.setClearButtonEnabled(True)
        
        choices_to_present = list(choices.values())
        qCompleterObj = QCompleter(QStringListModel(choices_to_present, self), self)
        qCompleterObj.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
        qCompleterObj.setFilterMode(Qt.MatchFlag.MatchContains)
        qCompleterObj.setCompletionMode(QCompleter.CompletionMode.PopupCompletion)
        
        self.setCompleter(qCompleterObj)
    # __ init__
    
    def selectedItem(self):
        """
        def selectedItem(self):
            returns the following dictionary: 
            return {'keys': [key for key, t in self.choices.items() if t==self.text()], 'text': self.text()}
            (all keys matching the text input)
        """
        return {'keys': [key for key, t in self.choices.items() if t==self.text()], 'text': self.text()}
    # selectedItem
    
    def addChoices(self, choices:Dict[Any, str]):
        self.choices.update(choices)
        
        choices_to_present = list(self.choices.values())
        newmodel = QStringListModel(choices_to_present, self)
        self.completer().setModel(newmodel)
    # addChoices


class cDictModel(QAbstractTableModel):
    def __init__(self, data, parent=None):
        """
        Initialize the model with a dictionary.
        
        Args:
            data (dict): Dictionary to model.
        """
        super().__init__(parent)
        self._data = data
        self._keys = list(data.keys())

    def rowCount(self, parent=None):
        """Return the number of rows in the model."""
        return len(self._data)

    def columnCount(self, parent=None):
        """Return the number of columns in the model (Key and Value)."""
        return 2  # One for the key and one for the value

    def data(self, index, role=Qt.DisplayRole):
        """Return the data for a given cell."""
        if not index.isValid() or role != Qt.DisplayRole:
            return None

        row = index.row()
        col = index.column()

        key = self._keys[row]
        if col == 0:  # Key column
            return key
        elif col == 1:  # Value column
            return self._data[key]
        return None

    def headerData(self, section, orientation, role=Qt.DisplayRole):
        """Return the header labels for rows or columns."""
        if role != Qt.DisplayRole:
            return None

        if orientation == Qt.Horizontal:
            return ["Key", "Value"][section]  # Column headers
        elif orientation == Qt.Vertical:
            return str(section + 1)  # Row numbers
        return None

    def setData(self, index, value, role=Qt.EditRole):
        """Set the data for a given cell."""
        if not index.isValid() or role != Qt.EditRole:
            return False

        row = index.row()
        col = index.column()
        key = self._keys[row]

        if col == 1:  # Only allow editing the value column
            self._data[key] = value
            self.dataChanged.emit(index, index, [role])
            return True

        return False

    def flags(self, index):
        """Set flags for each cell."""
        if not index.isValid():
            return Qt.NoItemFlags

        flags = Qt.ItemIsSelectable | Qt.ItemIsEnabled
        if index.column() == 1:  # Only value column is editable
            flags |= Qt.ItemIsEditable

        return flags

class cComboBoxFromDict(QComboBox):
    """
    Generates QComboBox from dictionary

    Args:
        dict (Dict): The input dictionary. The values will be the data returned by currentData, and
            the keys will be the values shown in the QComboBox
        parent (QWidget) default None
    
    """
    _combolist:List = []
    
    def __init__(self, dict:Dict, parent:QWidget = None):
        super().__init__(parent)
        
        # don't do completers - assume underlying QComboBox is non-editable
        # choices_to_present = list(dict)
        # qCompleterObj = QCompleter(QStringListModel(choices_to_present, self), self)
        # qCompleterObj.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
        # qCompleterObj.setFilterMode(Qt.MatchFlag.MatchContains)
        # qCompleterObj.setCompletionMode(QCompleter.CompletionMode.PopupCompletion)
        # self.setCompleter(qCompleterObj)
        
        self.replaceDict(dict)

    def replaceDict(self, dict:Dict[str, Any]):
        self.clear()
        self._combolist.clear()
        if isinstance(dict,Dict):
            for key, val in dict.items():
                self.addItem(key,val)
                self._combolist.append({key:val})

#########################################        
#########################################        

# cQRecordsetView - Scrollable layout of records
class cQRecordsetView(QWidget):
    _newdgt_fn:Callable[[], QWidget] = None
    _btnAdd:QPushButton = None
    def __init__(self, newwidget_fn:Callable[[], QWidget] = None, parent=None):
        """
        Widget which displays a set of records

        Args:
            newwidget_fn (Callable, optional): . Defaults to None. newwidget_fn() should return a new record 
                in a widget suitable for adding to this layout
            parent (_type_, optional): _description_. Defaults to None.
        """
        super().__init__(parent)
        self._newdgt_fn = newwidget_fn
        self.init_ui()

    def init_ui(self):
        self.mainLayout = QVBoxLayout(self)

        # set up scroll area
        self.scrollarea = QScrollArea(self)
        self.scrollarea.setWidgetResizable(True)
        self.scrollarea.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.scrollarea.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.mainLayout.addWidget(self.scrollarea)
    
        # Container widget for the layout
        self.scrollwidget = QWidget()
        # layout for the scrollwidget
        self.scrolllayout = QVBoxLayout(self.scrollwidget)
        # # Set container as the scroll area widget
        self.scrollarea.setWidget(self.scrollwidget)

        if self._newdgt_fn:
            self._btnAdd = QPushButton(self.scrollwidget)
            self._btnAdd.setObjectName('AddBtnQRS')
            self._btnAdd.setText('\nAdd\n')
            self._btnAdd.clicked.connect(self.addBtnClicked)
            self.scrolllayout.addWidget(self._btnAdd)

        self.init_recSet()
    # init_ui

    def setAddText(self, addText:str = '\nAdd\n'):
        ...
    # setAddText

    def addWidget(self, wdgt:QWidget):
        insAt = self.scrolllayout.count()-1 if self._newdgt_fn else -1
        self.scrolllayout.insertWidget(insAt, wdgt)
        line = QFrame(self)
        myWidth = self.geometry().width()
        line.setGeometry(QRect(0, 0, myWidth, 3))
        line.setFrameShape(QFrame.Shape.HLine)
        line.setFrameShadow(QFrame.Shadow.Sunken)
        self.scrolllayout.insertWidget(insAt+1, line)
    # addWidget

    # addLayout needed?

    def init_recSet(self):
        # remove all widgets from scrolllayout
        # for wdgt in self.scrolllayout:
        idx = 0
        child = self.scrolllayout.itemAt(idx)
        while child != None:
            if child.widget() == self._btnAdd:   # don't removew the Add Button!
                idx += 1
            else:
                widg = self.scrolllayout.takeAt(idx)
                widg.widget().deleteLater() # del the widget
            # endif child == self._btnAdd
            child = self.scrolllayout.itemAt(idx)
    # init_recSet
    
    @Slot()
    def addBtnClicked(self):
        if callable(self._newdgt_fn):
            self.addWidget(self._newdgt_fn())
    # addBtnClicked

#########################################        
#########################################        

def Excelfile_fromqs(qset:QSqlQueryModel|List[Dict[str, Any]], flName:str|None = None, 
                     freezecols:int = 0, returnFileName: bool = False) -> Workbook|str:
    """
    qset: a QAbstractTableModel or list of dictionaries
    flName: the name of the file to be built (WITHOUT extension!).  It's stored on the server.  If it's to be dl'd, the caller does that
    freezecols = 0: the number of columns to freeze to the left
    The top row contains the field names, is always frozen, is bold and is shaded grey

    used to Return the name of the Workbook file (with extension).  Once I start building in errorchecking and exceptions, other returns may be possible
    Returns the Workbook file (with extension)
    """

    # far easier to process a list of dictionaries, so...
    if isinstance(qset,QAbstractTableModel):
        # make this a util
        qlist = [ {qset.record(row).field(col).name() : qset.record(row).field(col).value() for col in qset.columnCount() } for row in range(qset.rowCount()) ]
    elif isinstance(qset,list):
        qlist = qset
    else:
        return None
    if qlist:
        if not isinstance(qlist[0],dict):
            # review this later ...
            try:
                qlist = [n.__dict__ for n in qlist]
            except:
                qlist = []

    # create empty workbook with an empty worksheet
    wb = Workbook()
    ws = wb.active

    # header row is names of columns
    if qlist:
        fields = list(qlist[0])
        ws.append(fields)

        # append each row
        for row in qlist:
            ws.append(list(row.values()))

        # make header row bold, shade it grey, freeze it
        # ws.show_gridlines = True  #Nope - this is a R/O attribute
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(fill_type=fills.FILL_SOLID,
                            start_color=colors.Color("00808080"),
                            end_color=colors.Color("00808080")
                            )
        #TODO: convert row1 and cols:freezecols to an address (A=0, B=1, C=2 etc) for line below
        ws.freeze_panes ='A2'
        #TODO: if freezecols passed, freeze them, too


    # save the workbook
    if flName:
        wb.save(flName+ExcelWorkbook_fileext)
    
    if returnFileName:
        # close the workbook
        wb.close()
        # and return file Name to the caller
        return flName+ExcelWorkbook_fileext
    else:
        # return the workbook itself
        return wb
    #endif returnFileName


class UpldSprdsheet():
    TargetModel = None
    SprdsheetDateEpoch = WINDOWS_EPOCH

    def SprdsheetFldDescriptor_creator(self, ModelFldName, AllowedTypes):
        """
        ModelFldName (str): the name of the field in the TargetModel
        AllowedTypes: list of tuples (type, cleanproc).  empty list if any string allowed
        """
        return  {
            # 'SprdsheetName': None,    # nope, this will be the index of SprdsheetFlds
            'ModelFldName': ModelFldName,
            'AllowedTypes': AllowedTypes,     
        }
    
    SprdsheetFlds = {}  # key will be the SprdsheetName, value is a SprdsheetFldDescriptor

    def cleanupfld(self, fld, val):
        usefld = False
        cleanval = None
        
        if fld not in self.SprdsheetFlds:
            # just feed the value back
            usefld = True
            cleanval = val
        elif not self.SprdsheetFlds[fld]['AllowedTypes']:
            usefld = (val is not None)
            if usefld: cleanval = str(val)
        else:
            for type, cleanproc in self.SprdsheetFlds[fld]['AllowedTypes']:
                if isinstance(val, type):
                    usefld = True
                    cleanval = cleanproc(val)
                    break
                #endif instance(val, type)
            #endfor type, cleanproc
        #endif fld not in self.SprdsheetFlds

    def process_spreadsheet(self, SprsheetName):
        pass


class QRawSQLTableModel(QAbstractTableModel):
    def __init__(self, rows:List[Dict[str,Any]], colNames:List[str], parent:QObject = None):
        super().__init__(parent)
        self.headers = colNames
        self.queryset = rows
    
    def rowCount(self, parent = QModelIndex()):
        return len(self.queryset)
    
    def columnCount(self, parent = QModelIndex()):
        return len(self.headers)
    
    def data(self, index, role = Qt.DisplayRole):
        if not index.isValid():
            return None
        if role == Qt.ItemDataRole.DisplayRole:
            rec = self.queryset[index.row()]
            fldName = self.headers[index.column()]
            value = rec[fldName]

            return value
        # endif role
        return None

    # not editable - don't need setData

    def headerData(self, section, orientation, role = Qt.DisplayRole):
        if role == Qt.DisplayRole:
            if orientation == Qt.Orientation.Horizontal:
                return self.headers[section]
            elif orientation == Qt.Orientation.Vertical:
                return str(section+1)
            #endif orientation
        # endif role
        return None


###################################################
###################################################
###################################################


class UnderConstruction_Dialog(QDialog):
    _svg_constr_barrier = 'assets/svg/under-construction-barrier-icon.svg'
    def __init__(self, parent:QWidget = None, constructionMsg:str = '', f:Qt.WindowType = Qt.WindowType.Dialog):
        super().__init__(parent, f)

        if not self.objectName():
            self.setObjectName(u"Dialog")
        self.resize(std_popdialogsize)
        self.setWindowTitle('Not Built Yet')
        self.buttonBox = QDialogButtonBox(self)
        self.buttonBox.setObjectName(u"buttonBox")
        self.buttonBox.setGeometry(QRect(30, 260, 341, 32))
        self.buttonBox.setOrientation(Qt.Orientation.Horizontal)
        self.buttonBox.setStandardButtons(QDialogButtonBox.StandardButton.Ok)
        self.buttonBox.setCenterButtons(True)
        self.constrsign = QSvgWidget(self._svg_constr_barrier,self)
        self.constrsign.setObjectName(u"constrwidget")
        self.constrsign.setGeometry(QRect(10, 60, 381, 191))
        self.label = QLabel(self)
        self.label.setObjectName(u"label")
        self.label.setGeometry(QRect(10, 10, 381, 51))
        font = QFont()
        font.setPointSize(12)
        font.setKerning(False)
        self.label.setFont(font)
        self.label.setAlignment(Qt.AlignmentFlag.AlignLeading|Qt.AlignmentFlag.AlignLeft|Qt.AlignmentFlag.AlignTop)
        self.label.setWordWrap(True)
        self.label.setText(constructionMsg)

        self.buttonBox.rejected.connect(self.reject)
        self.buttonBox.accepted.connect(self.accept)

        QMetaObject.connectSlotsByName(self)
    # __init__


class cQFmFldWidg(QWidget):
    _wdgt:QWidget = None
    _label:QLabel = None
    _labelSetLblText:Callable = None
    _modlField:str = None
    _lblChkYN:QLineEdit = None
    _lblChkYNValues:Dict[bool,str]|None = None

    signalFldChanged:Signal = Signal(object)

    def __init__(self, 
        widgType:type[QWidget], 
        lblText:str = ' ', 
        lblChkBxYesNo:Dict[bool,str]|None = None,
        alignlblText:Qt.AlignmentFlag = Qt.AlignmentFlag.AlignLeft ,
        modlFld:str = None, 
        choices:Dict|List = None, initval:str = '',
        parent:QWidget = None, 
        ):
        
        super().__init__(parent)

        self._wdgt = self.createWidget(widgType, choices, initval)
        
        lblText = self.tr(lblText)
        wdgt = self._wdgt
        
        # set hooks
        # if any([wdgt.inherits(tp) for tp in ['cDataList', ]]):
        if issubclass(widgType, (cDataList, )):
            self._label = QLabel(lblText)
            self.LabelText = self._label.text
            self._labelSetLblText = self._label.setText
            self._label.setBuddy(wdgt)

            self.Value    = wdgt.selectedItem
            self.setValue = wdgt.setText
            
            self.addChoices = wdgt.addChoices

            wdgt.editingFinished.connect(self.fldChanged)
        # elif any([wdgt.inherits(tp) for tp in ['QLineEdit', ]]):
        elif issubclass(widgType, (QLineEdit, )):
            self._label = QLabel(lblText)
            self.LabelText = self._label.text
            self._labelSetLblText = self._label.setText
            self._label.setBuddy(wdgt)

            self.Value    = wdgt.text
            self.setValue = wdgt.setText

            wdgt.editingFinished.connect(self.fldChanged)
        # elif any([wdgt.inherits(tp) for tp in ['QTextEdit', 'QPlainTextEdit' ]]):
        elif issubclass(widgType, (QTextEdit, QPlainTextEdit, )):
            self._label = QLabel(lblText)
            self.LabelText = self._label.text
            self._labelSetLblText = self._label.setText
            self._label.setBuddy(wdgt)

            self.Value    = wdgt.toPlainText
            self.setValue = wdgt.setPlainText

            wdgt.textChanged.connect(self.fldChanged)
        # elif any([wdgt.inherits(tp) for tp in ['cComboBoxFromDict', 'QComboBox', ]]):
        elif issubclass(widgType, (cComboBoxFromDict, QComboBox, )):
            self._label = QLabel(lblText)
            self.LabelText = self._label.text
            self._labelSetLblText = self._label.setText
            self._label.setBuddy(wdgt)

            self.Value    = wdgt.currentData
            self.setValue = lambda value: \
                wdgt.setCurrentText(str(value)) if wdgt.findData(value) == -1 else wdgt.setCurrentIndex(wdgt.findData(value))

            if isinstance(wdgt, cComboBoxFromDict):
                self.replaceDict = wdgt.replaceDict

            wdgt.activated.connect(self.fldChanged)
        # elif any([wdgt.inherits(tp) for tp in ['QDateEdit', ]]):
        elif issubclass(widgType, (QDateEdit, )):
            self._label = QLabel(lblText)
            self.LabelText = self._label.text
            self._labelSetLblText = self._label.setText
            self._label.setBuddy(wdgt)

            self.Value    = lambda: wdgt.date().toPython()
            self.setValue = wdgt.setDate

            wdgt.userDateChanged.connect(self.fldChanged)
        # support QSpinBox, QDoubleSpinBox, QDateTimeEdit, QTimeEdit ???
        # support QCalendarWidget later
        # elif any([self.inherits(tp) for tp in ['QCalendarWidget', ]]):
        # elif any([wdgt.inherits(tp) for tp in ['QCheckBox', ]]):
        elif issubclass(widgType, (QCheckBox, )):
            self._label = wdgt
            wdgt.setText(lblText)
            self.LabelText = wdgt.text
            self._labelSetLblText = wdgt.setText

            self.Value    = wdgt.isChecked
            self.setValue = lambda value: wdgt.setChecked(value if isinstance(value,bool) else False)
            
            if lblChkBxYesNo:
                self._lblChkYNValues = lblChkBxYesNo
                self._lblChkYN:QLineEdit = QLineEdit()
                self._lblChkYN.setProperty('noedit', True)
                self._lblChkYN.setReadOnly(True)
                self._lblChkYN.setFrame(False)
                self._lblChkYN.setMaximumWidth(40)
                self._lblChkYN.setFocusPolicy(Qt.FocusPolicy.NoFocus)

            wdgt.checkStateChanged.connect(self.fldChanged)
        # support later: QButtonGroup/QGroupBox
        # support later: QSlider, QDial
        else:
            raise TypeError(f'type {widgType} is not implemented')
        # endif widget type test

        # set the ModelField
        self.setModelField(modlFld)
        
        # set up the layout
        layout = QGridLayout()
        #(lblTupl(row,col),wdgtTupl(row,col))
        if alignlblText == Qt.AlignmentFlag.AlignLeft:
            positions = ((0,0),(0,1))
        elif alignlblText == Qt.AlignmentFlag.AlignRight:
            positions = ((0,1),(0,0))
        elif alignlblText == Qt.AlignmentFlag.AlignTop:
            positions = ((0,0),(1,0))
        elif alignlblText == Qt.AlignmentFlag.AlignBottom:
            positions = ((1,0),(0,0))
        else:
            # default to left
            positions = ((0,0),(0,1))
        
        # place widgets in layout
        if any([wdgt.inherits(tp) for tp in ['QCheckBox', ]]):
            if lblChkBxYesNo:
                layout.addWidget(self._lblChkYN,*positions[0])
                layout.addWidget(wdgt,*positions[1])
            else:
                layout.addWidget(wdgt,0,0)
        else:
            if lblText:
                layout.addWidget(self._label,*positions[0])
                layout.addWidget(wdgt,*positions[1])
            else:
                layout.addWidget(wdgt,0,0)
        #endif a checkbox
        self.setLayout(layout)

    # this can be overridden in case there are "superclasses"
    def createWidget(self, widgType:type[QWidget], choices:Dict|List = None, initval:str = '',) -> QWidget:
        wdgt = None
        if issubclass(widgType,(cComboBoxFromDict, )):
            wdgt = widgType(choices, self)
        elif issubclass(widgType, (cDataList, )):
            wdgt = widgType(choices, initval, self)
        elif issubclass(widgType, (QComboBox, )):
            # don't use this widget if using a model, or
            # clear(), then addItem()
            wdgt = widgType(self)
            wdgt.addItems(choices)  
        else:
            wdgt = widgType(self)  # does this have to be by type?
        #endif widgType 
            
        return wdgt
    #createWidget

    def setLabelText(self, txt:str) -> None:
        self._labelSetLblText(txt)
        self._label.repaint()
        
    def __getattr__(self, name):
        # Delegate attribute access to the contained widget if the attribute
        # is not found in the cQF instance.
        return getattr(self._wdgt, name)

    def modelField(self) -> str:
        return self._modlField
    def setModelField(self, fldName:str) -> None:
        self._modlField = fldName

    @Slot()
    def fldChanged(self, *args):
        if self._lblChkYNValues:
            # changeit
            newstate = (args[0] == Qt.CheckState.Checked)
            self._lblChkYN.setText(self._lblChkYNValues[newstate])
        self.signalFldChanged.emit(args if args else (None,))


class cQFmNameLabel(QLabel):
    def __init__(self, formName:str = '', parent:QWidget = None):
        super().__init__(parent)
        
        fontFormTitle = QFont()
        fontFormTitle.setFamilies([u"Copperplate Gothic"])
        fontFormTitle.setPointSize(24)
        self.setFont(fontFormTitle)
        self.setFrameShape(QFrame.Shape.Panel)
        self.setFrameShadow(QFrame.Shadow.Raised)
        self.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.setWordWrap(True)
        
        if formName:
            self.setText(formName)
        

##################################################
##################################################
##################################################


class cQSqlTableModel(QSqlTableModel):
    retListofQSQLRecord = -1
    def __init__(self, tblName:str, db:QSqlDatabase = QSqlDatabase.database(), parent:QObject = None):
        super().__init__(parent, db)
        self.setTable(tblName)
        # why does setEditStrategy crash?
        self.setEditStrategy(QSqlTableModel.EditStrategy.OnManualSubmit)    # think about this - pass as parm?
        self.select()
    
    def recordsetList(self, retFlds:int|List[str] = retListofQSQLRecord, filter:str = None) -> List:
        retList = []
        if filter:
            self.setFilter(filter)
        self.select()
        if retFlds == '*' or (isinstance(retFlds,List) and retFlds[0]=='*'):
            retFlds = [rec.fieldName(i) for i in range(rec.count())]
        for n in range(self.rowCount()):
            rec = self.record(n)
            if retFlds == self.retListofQSQLRecord:
                retList.append(rec)
            elif isinstance(retFlds,List):
                retList.append({fldName:rec.value(fldName) for fldName in retFlds})
            else:
                raise f'Invalid Field List {retFlds}'
            #endif retFlds
        #endwhile not rec.isEmpty()
        
        return retList
    #enddef recordsetList
class cQSqlRelationalTableModel(QSqlRelationalTableModel):
    retListofQSQLRecord = -1
    def __init__(self, tblName:str, db:QSqlDatabase = QSqlDatabase.database(), parent:QObject = None):
        super().__init__(parent, db)
        self.setTable(tblName)
        # why does setEditStrategy crash?
        self.setEditStrategy(QSqlTableModel.EditStrategy.OnManualSubmit)    # think about this - pass as parm?
        self.select()
    
    def recordsetList(self, retFlds:int|List[str] = retListofQSQLRecord, filter:str = None) -> List:
        retList = []
        if filter:
            self.setFilter(filter)
        self.select()
        if retFlds == '*' or (isinstance(retFlds,List) and retFlds[0]=='*'):
            retFlds = [rec.fieldName(i) for i in range(rec.count())]
        for n in range(self.rowCount()):
            rec = self.record(n)
            if retFlds == self.retListofQSQLRecord:
                retList.append(rec)
            elif isinstance(retFlds,List):
                retList.append({fldName:rec.value(fldName) for fldName in retFlds})
            else:
                raise f'Invalid Field List {retFlds}'
            #endif retFlds
        #endwhile not rec.isEmpty()
        
        return retList
    #enddef recordsetList


##################################################
##################################################
##################################################

import ast
dividerchar = '\u23FA'
def show_fns(path_:str):
    # open file as ast (abstract syntax tree)
    with open(path_) as file:
        node = ast.parse(file.read())

    # 
    def show_info(functionNode:ast.FunctionDef|ast.ClassDef):
        function_rep = ''
        function_rep = functionNode.name + '('

        for arg in functionNode.args.args:
            function_rep += arg.arg + ','

        function_rep = function_rep.rstrip(function_rep[-1])
        rNode = functionNode.returns
        rtype = None
        if isinstance(rNode, ast.BinOp):
            rtype = f'{rNode.left.id} | {rNode.right.id}'
        elif isinstance(rNode, ast.Subscript):
            rtype = f'{rNode.value.id}'
        elif isinstance(rNode, ast.Attribute):
            rtype = f'{rNode.value.id}'
        elif isinstance(rNode, ast.Constant):
            rtype = f'{rNode.value}'
        elif isinstance(rNode, ast.Name):
            rtype = f'{rNode.id}'
        #endif rNode tyupe
        function_rep += f') -> {rtype} {dividerchar} lines {functionNode.lineno} to {functionNode.end_lineno}'
        return function_rep

    # get all fns and classes
    result = {'classes':[], 'functions':[]}
    functions = [n for n in node.body if isinstance(n, ast.FunctionDef)]
    classes = [n for n in node.body if isinstance(n, ast.ClassDef)]

    for function in functions:
        result['functions'].append(f'def {show_info(function)}')

    for class_ in classes:
        result['classes'].append(f'class {class_.name}({[NN.id for NN in class_.bases]}) {dividerchar} lines {class_.lineno} to {class_.end_lineno}')
        methods = [n for n in class_.body if isinstance(n, ast.FunctionDef)]
        for method in methods:
            result['classes'].append(f'    def {class_.name}.{show_info(method)}')

    return result
    # print(', '.join(result))
    # This prints expected output
    # fo(x), A.fun(self,y), A._bo(self,y), A.NS(y,z), B.foo(self,z), B._bar(self,t)
def pretty_show_fns(path_:str):
    result = show_fns(path_)

    result_str = ''

    result_str += ' CLASSES: \n----------\n'
    for c in result['classes']:
        result_str += f'{c}\n'
    result_str += '\n'

    result_str += ' FUNCTIONS: \n------------\n'
    for c in result['functions']:
        result_str += f'{c}\n'
    
    return result_str
